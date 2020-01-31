# instagram_crawling_10post.py
# 뽑은 아이디로 각자의 프로필화면으로 들어가 게시글 사진 10개 다운받기.

import requests
import urllib.request
import random
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import os
import sys
non_bmp_map = dict.fromkeys(range(0x10000, sys.maxunicode + 1), 0xfffd)


# 엑셀 파일로 저장된 id(B1~B200) 있으면 그걸 arr_id 에 넣는다.
arr_profile = []                                                                        # 프로필 화면 주소들 담김.
arr_id = []                                                                             # 아이디 담김.

# 엑셀 파일 불러옴, id로 프로필 주소 만들기
# https://www.instagram.com/(id 넣음)/?hl=ko 형식
from openpyxl import load_workbook

wb = load_workbook("instagramWithForm.xlsx")
ws = wb.active
for i in range(2,202):
    id = ws["B"+str(i)].value
    profile = "https://www.instagram.com/"+ id +"/?hl=ko"
    print("[%d] : %s"%(i,id))
    arr_id.append(id)
    arr_profile.append(profile)
    
    
print("=================================================================")
print("id 수 : %d, 프로필 주소 수 : %d" %(len(arr_id), len(arr_profile)))


# 드라이버 연결
driver = webdriver.Chrome(executable_path="D:\\python_D\\chromedriver.exe")

count = 0
count_img = 0

# 4. 뽑은 아이디로 각자의 프로필화면으로 들어가 프로필 다운받기.
for profile in arr_profile:
    if(arr_id[count] != "none"):
        driver.implicitly_wait(2)
        driver.get(profile)     # 프로필화면으로 들어감

        # 프로필 들어가서 게시물 이미지 검색
        try:   
            href = driver.find_elements_by_xpath('//*[@class="KL4Bh"]')                       # <div class="KL48h"> <img src="주소"> 형식.
            print("게시물 클래스 검색성공", len(href))
            print("post[%d]" % count)
        except:
            print("게시물 클래스 검색실패")

        # 이미지 저장위한 폴더생성
        folder = ''
        try:
            folder = ".\\크롤러\\오늘의훈남2\\"+arr_id[count]
            os.mkdir(folder)                                                                     # 각 id에 대한 폴더 생성함.
            print("폴더 생성 성공")
        except:
            print(folder+": 폴더 생성 실패 또는 이미 만들어짐.")
        # 이미지 저장.
        for i in range(0,10):
            try:
                full_name = folder+"\\"+arr_id[count]+"_"+str(i)+".jpg"                            # folder 경로 밑에 "id_0.jpg" 형태로 저장됨.
                print("저장될 이미지 : ",full_name)
                href_profile = href[i].find_element_by_css_selector('img').get_attribute('src')      # 게시물 사진 검색
                print("href_profile : ", href_profile)
    #            urllib.request.urlretrieve(item.get_attribute('src'), full_name) # src를 받는다.
                urllib.request.urlretrieve(href_profile, full_name)                                  # full_name에 받아온 이미지 저장.
            
                print("img[%d] : %s" %(count, arr_id[count]+".jpg"))

                count_img += 1
            except:
                print("이미지 저장 실패 또는 이미지가 없음.")
                
                
            print("-------------------------")
        
    count += 1
    
driver.quit()
print("이미지 개수 : %d"%count_img)
print("Saved!")
