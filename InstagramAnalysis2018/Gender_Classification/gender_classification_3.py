# gender_classification_3.py
# 프로필사진 하나가 아닌 게시글 10개 사진에 대해 분석(10개중에 여성:5, 남성:3, 모름:2 => 여성)
# 더 많은 label 사용
# 게시글 사진 10개를 Google Vision API를 사용해 분석했을 때 나오는 레이블 값에 따라 성별 분류
#      ["girl", "beauty", "long hair", "woman"]라는 단어가 나올 경우 [여성(1)]
#      ["boy", "military", "suit", "man", "male"]라는 단어가 나올 경우 [남성(0)]
#      안나올 경우 [모름(9999)]

# run_quickstart("test.jpg") 로 부르면 실행됨.
def run_quickstart(file_name):
    import io
    import os
 
    # 구글 라이브러리 import
    from google.cloud import vision
    from google.cloud.vision import types
 
    # 사용할 클라이언트 설정
    client = vision.ImageAnnotatorClient()
    
    # 이미지 읽기
    with io.open(file_name, 'rb') as image_file:
        content = image_file.read()
 
    image = types.Image(content=content)
 
    # label 뽑아냄.
    response = client.label_detection(image=image)
    labels = response.label_annotations
 
    print('Labels:')
    for label in labels:
        labeltext = label.description
        if(labeltext=="girl" or labeltext=="beauty" or labeltext=="long hair" or labeltext=="woman"): # girl, beauty, long hair, woman 이라면 1리턴
            print(label.description + " = " + str(int(label.score*100)) + "%")
            return 1
        elif(labeltext=="boy" or labeltext=="military" or labeltext=="suit" or labeltext=="man" or labeltext=="male"):  # boy, military, suit, man, male 라면 0리턴
            print(label.description + " = " + str(int(label.score*100)) + "%")
            return 0

    return 9999   # 성별구분 못하면 9999리턴

# 엑셀 파일 불러옴, id 순서대로 다 불러오기.
from openpyxl import load_workbook

wb = load_workbook(".\\instagramWithForm.xlsx")
ws = wb.active

arr_id = []
for i in range(2,202):
    id = ws["B"+str(i)].value
#    print("[%d] : %s"%(i,id))
    arr_id.append(id)
    

print("=================================================================")
print("id 수 : %d, [B2]~[B202]" %(len(arr_id)))

# D:\python_D\크롤러\오늘의훈남2\id\(id.jpg) id별로 오늘의훈남2 밑의 id 폴더 경로에 사진이 저장됨.(최대 10개)
# 하나하나 들어가서 레이블값 받아오기.
filepath = ".\\크롤러\\오늘의훈남2\\"

ws2 = wb.active
count = 2
countgender = 0
for id in arr_id:                           # id 별로
    girl = 0                                # 여자라고 나온 횟수
    boy = 0                                 # 남자라고 나온 횟수
    none = 0                                # 성별보른 횟수
    
    for i in range(0,10):                   # 10개 이하 사진 있음.
        filename = filepath+id+"\\"+id+"_"+str(i)+".jpg"   # 오늘의훈남2\\id\\id_0.jpg ~ id_9.jpg 까지 있음.
        print("[%d] %s"%(count, filename))
        
        try:
            gender = run_quickstart(filename)
            if(gender == 0):        # 0 : 남
                boy += 1
            elif(gender == 1):      # 1 : 여
                girl += 1
            else:                   # 9999 : 성별없음.
                none += 1         
#            print("성별 구분함.")
        except:
            print("사진이 없음.")
#            ws2.cell(row=count, column=1, value=int(9999)) # "A"의 2번째부터 적어줌.

    if(girl > boy):
        gender = 1
        countgender += 1
    elif(girl < boy):
        gender = 0
        countgender += 1
    else:
        gender = 9999

    print("girl = %d, boy = %d"%(girl, boy))
    ws2.cell(row=count, column=1, value=int(gender)) # "A"의 2번째부터 적어줌.(0 : 남, 1 : 여, 9999 : 모름)
    ws2.cell(row=count, column=5, value=str(boy)+":"+str(girl))
    count += 1
    print("======================================")
wb.save(".\\instagramWithForm.xlsx") # 저장

print("성별 구분된 id : %d"%countgender)
