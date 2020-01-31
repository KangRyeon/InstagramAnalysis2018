# gender_classification_1.py
# 프로필 사진을 Google Vision API를 사용해 분석했을 때 나오는 레이블 값에 따라 성별 분류
#      ["girl", "beauty"]라는 단어가 나올 경우 [여성(1)]
#      ["boy"]라는 단어가 나올 경우 [남성(0)]
#      안나올 경우 [모름(9999)]

# run_quickstart("test.jpg") 로 부르면 실행됨.
def run_quickstart(file_name):
    import io
    import os
 
    # 구글 라이브러리 import
    from google.cloud import vision
    from google.cloud.vision import types
 
    # 사용할 클라이언트 설정
    client = vision.ImageAnnotatorClient()
    
    # 이미지 읽기
    with io.open(file_name, 'rb') as image_file:
        content = image_file.read()
 
    image = types.Image(content=content)
 
    # label 뽑아냄.
    response = client.label_detection(image=image)
    labels = response.label_annotations
 
    print('Labels:')
    for label in labels:
        if(label.description=="girl" or label=="beauty"): # girl 또는 beauty 라면 1리턴
            print(label.description + " = " + str(int(label.score*100)) + "%")
            return 1
        elif(label.description=="boy"):  # boy 라면 0리턴
            print(label.description + " = " + str(int(label.score*100)) + "%")
            return 0
    return 9999   # 성별구분 못하면 9999리턴



# 엑셀 파일 불러옴, id 순서대로 다 불러오기.
from openpyxl import load_workbook

wb = load_workbook("instagramWithForm.xlsx")
ws = wb.active

arr_id = []
for i in range(2,202):
    id = ws["B"+str(i)].value
#    print("[%d] : %s"%(i,id))
    arr_id.append(id)
    

print("=================================================================")
print("id 수 : %d, [B2]~[B202]" %(len(arr_id)))

# .\크롤러\오늘의훈남\(id.jpg) id별로 오늘의훈남 경로에 사진이 저장되어 있음.
# 각 사진마다 불러와 레이블값 받아오기.
filepath = ".\\크롤러\\오늘의훈남2\\"

ws2 = wb.active
count = 2
countgender = 0
for id in arr_id:
    filename = filepath+id+".jpg"
    print("[%d] %s"%(count, filename))
    try:
        gender = run_quickstart(filename)
        if(gender != 9999):                  # 성별 구분했다면(9999 = 성별모름)
            ws2.cell(row=count, column=1, value=int(gender)) # "A"의 2번째부터 적어줌.(0 : 남, 1 : 여)
            print("성별 구분함.")
            countgender += 1
    except:
        print("사진이 없음.")
        ws2.cell(row=count, column=1, value=int(9999)) # "A"의 2번째부터 적어줌.
    count += 1
    
wb.save(".\\instagramWithForm.xlsx") # 저장

print("성별 구분된 id : %d"%countgender)
