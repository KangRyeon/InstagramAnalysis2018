# gender_label_find.py
# 성별 구분한 label 값을 찾아내기 위해 [girl, beauty], [boy]와
# 함께 나온 label이 무엇인지 파악해 성별분류 가능한 label값 찾아내기

# run_quickstart("test.jpg") 로 부르면 실행됨.
from openpyxl import load_workbook

def run_quickstart(file_name):
    import io
    import os
 
    # 구글 라이브러리 import
    from google.cloud import vision
    from google.cloud.vision import types
 
    # 사용할 클라이언트 설정
    client = vision.ImageAnnotatorClient()
    
    # 이미지 읽기
    with io.open(file_name, 'rb') as image_file:
        content = image_file.read()
 
    image = types.Image(content=content)
 
    # label 뽑아냄.
    response = client.label_detection(image=image)
    labels = response.label_annotations

    # 성별구분된 레이블 값들 저장위해 파일을 열기.   
    print('Labels:')
    for label in labels:
        
        if(label.description=="girl" or label=="beauty"): # girl 또는 beauty 라면 1리턴
            print(label.description + " = " + str(int(label.score*100)) + "%")
            for gender in labels:
                findLabel(1, gender.description)
            return 1
        elif(label.description=="boy"):  # boy 라면 0리턴
            print(label.description + " = " + str(int(label.score*100)) + "%")
            for gender in labels:
                findLabel(0, gender.description)
            return 0

    
    return 9999   # 성별구분 못하면 9999리턴

# instagramGenderLabel.xlsx 에서 성별구분하기, gender(0 또는 1), label("girl",...)
def findLabel(gender, label):
    print("gender = %d, label = %s"%(gender, label))
    # 엑셀 파일 불러오기
    wb1 = load_workbook(".\\instagramGenderLabel.xlsx")
    ws1 = wb1.active

    # 남자일때 = B2~B50 확인해 label값 있으면 C 열 값 받아와 +1 해줌
    # 없으면 label 추가, C 열 값 1로 초기화
    if(gender == 0):
        print("남자임")
        for i in range(2, 101):   # 2~100
            text = ws1["B"+str(i)].value
            if(text == label): # label 값이 있으면
                cnt = ws1["C"+str(i)].value
                ws1.cell(row=i, column=3, value=int(cnt)+1)
                break;
            if(text == None): # 아무것도 없는 칸이 나오면
                ws1.cell(row=i, column=1, value=int(0))
                ws1.cell(row=i, column=2, value=label)
                ws1.cell(row=i, column=3, value=int(1))
                break;
                
    # 여자일때    
    elif(gender == 1):
        print("여자임")
        for i in range(101, 201):   # 101~200
            text = ws1["B"+str(i)].value
            if(text == label): # label 값이 있으면
                cnt = ws1["C"+str(i)].value
                ws1.cell(row=i, column=3, value=int(cnt)+1)
                break;
            if(text == None): # 아무것도 없는 칸이 나오면
                ws1.cell(row=i, column=1, value=int(1))
                ws1.cell(row=i, column=2, value=label)
                ws1.cell(row=i, column=3, value=int(1))
                break;
    wb1.save("instagramGenderLabel.xlsx")
    
# 엑셀 파일 불러옴, id 순서대로 다 불러오기.

wb = load_workbook("instagramWithForm.xlsx")
ws = wb.active

arr_id = []
for i in range(2,202):
    id = ws["B"+str(i)].value
    arr_id.append(id)

print("=================================================================")
print("id 수 : %d, [B2]~[B202]" %(len(arr_id)))

# .\크롤러\오늘의훈남2\id\(id.jpg) id별로 오늘의훈남2 밑의 id 폴더 경로에 사진이 저장됨.(최대 10개)
# 하나하나 들어가서 레이블값 받아오기.
filepath = ".\\크롤러\\오늘의훈남2\\"

count = 2
countgender = 0
for id in arr_id:                           # id 별로
    girl = 0                                # 여자라고 나온 횟수
    boy = 0                                 # 남자라고 나온 횟수
    none = 0                                # 성별보른 횟수
    
    for i in range(0,10):                   # 10개 이하 사진 있음.
        filename = filepath+id+"\\"+id+"_"+str(i)+".jpg"   # 오늘의훈남2\\id\\id_0.jpg ~ id_9.jpg 까지 있음.
        print("[%d] %s"%(count, filename))
        
        try:
            gender = run_quickstart(filename)
            if(gender == 0):        # 0 : 남
                boy += 1
            elif(gender == 1):      # 1 : 여
                girl += 1
            else:                   # 9999 : 성별없음.
                none += 1         
        except:
            print("사진이 없음.")
            break;


    if(girl > boy):
        gender = 1
        countgender += 1
    elif(girl < boy):
        gender = 0
        countgender += 1
    else:
        gender = 9999

    print("girl = %d, boy = %d"%(girl, boy))
    count += 1
    print("======================================")

print("성별 구분된 id : %d"%countgender)
