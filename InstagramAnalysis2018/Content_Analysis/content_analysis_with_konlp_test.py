# content_analysis_with_konlp_test.py
# CONTENT 형태소 뽑아내기 완료
from collections import namedtuple
from gensim.models import doc2vec
from konlpy.tag import Twitter
import multiprocessing
from pprint import pprint
from openpyxl import load_workbook
import sys

non_bmp_map = dict.fromkeys(range(0x10000, sys.maxunicode + 1), 0xfffd)

twitter = Twitter()

def read_data(filename):
    wb = load_workbook(".\\instagramWithForm.xlsx")
    ws = wb.active
    # 젠더, 아이디, 내용, 태그, 젠더
    data = [[ws["A"+str(i)].value, ws["B"+str(i)].value , ws["C"+str(i)].value, ws["D"+str(i)].value, ws["F"+str(i)].value] for i in range(2,202)]
    # 내용, 태그
    # data = [[ws["C"+str(i)].value, ws["D"+str(i)].value] for i in range(2,202)]
    return data 
    '''
    with open(filename, 'r', encoding='utf-8') as f:
        data = [line.split('\t') for line in f.read().splitlines()]
    return data
    '''
def tokenize(doc, token):
  # norm, stem은 optional
  return [token.join(t) for t in twitter.pos(doc, norm=True, stem=True)]
  #return [t for t in twitter.pos(doc, norm=True, stem=True)] -> 글만뽑기
#doc2vec parameters
cores = multiprocessing.cpu_count()

vector_size = 300
window_size = 15
word_min_count = 2
sampling_threshold = 1e-5
negative_size = 5
train_epoch = 100
dm = 1
worker_count = cores

# 트레이닝 데이터 읽기
train_data = read_data('instagramWithForm.xlsx')
'''
for i in range(0,5):
    try:
        print("train data : %s, %s"%(train_data[i][2], train_data[i][3]))
    except:
        print("이모티콘임")
        print("train data : %s, %s"%(train_data[i][2].translate(non_bmp_map).encode("utf-8"), train_data[i][3]))
    print("=============================")
'''
print("train_data = ",len(train_data))

# 형태소 분류(문장에서 낱말뽑아냄.)
'''
for row in train_data[0:5]:
    train_docs = [(tokenize(row[2], "/"), row[3])]
    try:
        print(row[2])
    except:
        print(row[2].translate(non_bmp_map).encode("utf-8"))
'''
    #number = [0,8,9,16,24,28,32,40,48,56]
    #number = [0,8,9,16,24,32,40,41,42]
    #number = [56,64,66,68,72,80,81,88,89]
    #number = [80,87,88,89,90]
number = [170,174,180,182,183,184,192,200]
    #number = [0,8,9,16,24,32,40,48,56,64,66,68,72,80,87,88,90,96,104,112,120,128,136,144,152,160,168,170,174,180,182,183,184,192,200]
print(len(number))
for excel in range(0,len(number)-1):
    train_docs = [(tokenize(row[2], "/"), row[3]) for row in train_data[number[excel]:number[excel+1]]]
    print("문장에서 낱말 뽑아냄 : %d ~ %d"%(number[excel], number[excel+1]))
#    print(train_docs[0][0][0][0])
    '''
    #print("train_docs : ",train_docs[0:5])
    for i in range(0,9):
        print("[%d]"%(i+2))
        try:
            print("train docs(content) : %s"%(train_docs[i][0]))
        except:
            print("이모티콘임")      # 이모티콘이라 그냥 이모티콘인 부분만 출력
            print("train docs(content) : %s"%(train_docs[i][0][0].translate(non_bmp_map).encode("utf-8")))
        try:
            print("train docs(tag) : %s"%(train_docs[i][1]))
        except:
            print("이모티콘임")
            print("train docs(tag) : %s"%(train_docs[i][1].translate(non_bmp_map).encode("utf-8")))
    
    # 형태소 분류한 것 저장하기
    wb = load_workbook("instagramWithForm.xlsx")
    ws = wb.active

    # i = 0~199 (총 CONTENT 개수)
    # j = 2~201 (엑셀에서 위치) = i+2
    # k = 0~19 (형태소로 분류된 낱말 개수)
    for i in range(0,number[excel+1]-number[excel]):
        text = ""
        try:
            for k in range(0,20):
                text += train_docs[i][0][k]+", "  # "," 으로 형태소뽑은것 구분
        except:
                print("[%d]형태소 ', '으로 구분"%(i))
        print("================================")
#        print("뽑은 형태소 : ",text.translate(non_bmp_map).encode("utf-8"))
        try:
            ws["G"+str(i+2+number[excel])] = text
            wb.save("instagramWithForm.xlsx")
        except:
            ws["G"+str(i+2+number[excel])] = "None"
            wb.save("instagramWithForm.xlsx")



'''




        
